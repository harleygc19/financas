import pandas as pd
from openpyxl import load_workbook
import streamlit as st


df = pd.read_excel('Contas.xlsx')


def formulario():

    try:
        df = pd.read_excel('Contas.xlsx')
        
        # Convertendo as colunas de datas para o formato dd/mm/aa
        colunas_de_data = ['Data_vencimento', 'Data_do_pagamento']  # Substitua pelos nomes corretos das suas colunas de data
        
        for coluna in colunas_de_data:
            df[coluna] = pd.to_datetime(df['coluna'], errors='coerce').dt.strftime('%d/%m/%y')
    
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Categoria', 'Conta', 'Data_vencimento', 'Data_do_pagamento', 'Status_do_pagamento', 'Valor_pago', 'OBS'])
    
    return df

def formulario():    
    lista_status = 'Pago','Pendente'
    lista_categoria = 'Cartão de crédico','Saúde', 'Moradia', 'Lazer', 'Transporte', 'Alimentação', 'Telefonia', 'Educação', 'Outros'
    lista_contas = 'Luz', 'Celular Harley', 'Fies', 'Celular Daiana', 'Cartão Bradesco', 'Cartão Nubank', 'Plano de saúde','Aluguel', 'Gás', 'Condomínio', 'Faculdade','Mercado', 'Internet', 'Academia Harley', 'Academia Daiana', 'Outros', 'Cartão Moises', 'Cartão Fábio','Açougue','IPTV'
    if ' Categoria' not in st.session_state:
        st.session_state['Categoria'] = ''
    if 'Conta' not in st.session_state:
        st.session_state['Conta'] = ''
    if ' Data vencimento' not in st.session_state:
        st.session_state['Data_vencimento'] = ''
    if 'Data do pagamento' not in st.session_state:
        st.session_state['Data_do_pagamento'] = ''
    if 'Status do pagamento' not in st.session_state:
        st.session_state['Status_do_pagamento'] = ''
    if 'Valor_pago' not in st.session_state:
        st.session_state['Valor_pago'] = ''
    if 'OBS' not in st.session_state:
        st.session_state['OBS'] = ''
    
  
    st.session_state['Categoria'] = st.selectbox('Categoria',lista_categoria)
    st.session_state['Conta'] = st.selectbox('Conta', lista_contas)
    st.session_state['Data_vencimento']=st.date_input('Data vencimento')
    st.session_state['Data_do_pagamento']=st.date_input('Data do pagamento')
    st.session_state['Status_do_pagamento'] = st.selectbox('Status do Pagamento',lista_status)
    st.session_state['Valor_pago']=st.number_input('Valor pago')
    st.session_state['OBS'] = st.text_input('OBS')



    if st.button('Cadastrar'):
        cadastrar1(
        st.session_state['Categoria'],
        st.session_state['Conta'],
        st.session_state['Data_vencimento'],
        st.session_state['Data_do_pagamento'],
        st.session_state['Status_do_pagamento'],
        st.session_state['Valor_pago'],
        st.session_state['OBS'])
        
        st.success('Cadastro enviado com sucesso!')
        
        

def cadastrar1(Categoria, Conta, Data_vencimento,  Data_do_pagamento, Status_do_pagamento, Valor_pago, OBS):

    nova_linha = {'Categoria': Categoria,
                  'Conta': Conta,
                  'Data_vencimento': Data_vencimento,
                  'Data_do_pagamento': Data_do_pagamento,
                  'Status_do_pagamento': Status_do_pagamento,
                  'Valor_pago': Valor_pago,
                  'OBS': OBS}
    df_nova_linha = pd.DataFrame([nova_linha])
    try:
        wb = load_workbook (filename="Contas.xlsx")
    except FileNotFoundError:
        df_nova_linha.to_excel("Contas.xlsx", index=False)
        wb = load_workbook(filename="Contas.xlsx")
    ws = wb.active
    proxima_linha = ws.max_row + 1

    for index, row in df_nova_linha.iterrows():
        for col, value in enumerate(row, start=1):
            ws.cell(row=proxima_linha, column=col, value=value)
    proxima_linha += 1

    wb.save(filename="Contas.xlsx")
    return df




def tela_formulario():
    
    formulario()
   

    

def tela_tabela():
    df = pd.read_excel('Contas.xlsx') 
    st.dataframe(df,hide_index=True)

